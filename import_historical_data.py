#!/usr/bin/env python3
"""
Script to extract historical cash data from DiarioZ.xlsx and import to database.
"""

import openpyxl
import sys

def clean_value(value):
    """Convert value to decimal"""
    if value is None or value == '':
        return '0.00'
    try:
        return f"{float(value):.2f}"
    except:
        return '0.00'

def extract_graficas_sheet(file_path):
    """Extract data from Graficas sheet"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['Graficas']
    
    print(f"Processing sheet: {sheet.title}")
    
    hostel_data = []
    tienda_data = []
    
    # Hostel section starts at row 2 (row 1 is header)
    # Tienda section starts at row 16 (row 15 is header)
    
    # Extract Hostel data (rows 2-13)
    for row_idx in range(2, 14):  # 2014-2025 = 12 years
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        year = int(row[0]) if row[0] else None
        
        if year and 2014 <= year <= 2025:
            # Extract monthly values
            # Structure: Year, Jan_Hostel, Jan_Tienda, Feb_Hostel, Feb_Tienda, ...
            # Columns: 0=Year, 1=Jan, 2=blank, 3=Feb, 4=blank, 5=Mar, ...
            monthly_values = []
            for month in range(1, 13):
                col_idx = 1 + (month - 1) * 2  # 1, 3, 5, 7, ...
                value = row[col_idx] if col_idx < len(row) else None
                cleaned_value = clean_value(value)
                monthly_values.append(cleaned_value)
            
            hostel_data.append((year, monthly_values))
            print(f"  Hostel {year}: Jan={monthly_values[0]}, Feb={monthly_values[1]}, Mar={monthly_values[2]}")
    
    # Extract Tienda data (rows 16-27)
    for row_idx in range(16, 28):  # 2014-2025 = 12 years
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        year = int(row[0]) if row[0] else None
        
        if year and 2014 <= year <= 2025:
            # Extract monthly values
            monthly_values = []
            for month in range(1, 13):
                col_idx = 1 + (month - 1) * 2  # 1, 3, 5, 7, ...
                value = row[col_idx] if col_idx < len(row) else None
                cleaned_value = clean_value(value)
                monthly_values.append(cleaned_value)
            
            tienda_data.append((year, monthly_values))
            print(f"  Tienda {year}: Jan={monthly_values[0]}, Feb={monthly_values[1]}, Mar={monthly_values[2]}")
    
    return hostel_data, tienda_data

def generate_sql_inserts(hostel_data, tienda_data):
    """Generate SQL INSERT statements"""
    sql_statements = []
    
    # Clear existing data
    sql_statements.append("DELETE FROM historical_cash_data;")
    
    # Insert hostel data
    for year, monthly_values in hostel_data:
        for month, total_z in enumerate(monthly_values, start=1):
            if float(total_z) > 0:  # Only insert non-zero values
                sql = f"INSERT INTO historical_cash_data (year, month, businessType, totalZ, totalCash, totalCards) VALUES ({year}, {month}, 'hostel', {total_z}, 0, 0);"
                sql_statements.append(sql)
    
    # Insert tienda data
    for year, monthly_values in tienda_data:
        for month, total_z in enumerate(monthly_values, start=1):
            if float(total_z) > 0:  # Only insert non-zero values
                sql = f"INSERT INTO historical_cash_data (year, month, businessType, totalZ, totalCash, totalCards) VALUES ({year}, {month}, 'tienda', {total_z}, 0, 0);"
                sql_statements.append(sql)
    
    return sql_statements

def main():
    file_path = '/home/ubuntu/upload/DiarioZ.xlsx'
    
    print("Extracting data from DiarioZ.xlsx...")
    hostel_data, tienda_data = extract_graficas_sheet(file_path)
    
    print(f"\nExtracted {len(hostel_data)} years of Hostel data")
    print(f"Extracted {len(tienda_data)} years of Tienda data")
    
    # Generate SQL
    sql_statements = generate_sql_inserts(hostel_data, tienda_data)
    
    # Save to file
    output_file = '/home/ubuntu/hostel_management_app/import_historical_data.sql'
    with open(output_file, 'w') as f:
        f.write('\n'.join(sql_statements))
    
    print(f"\nSQL statements saved to {output_file}")
    print(f"Total statements: {len(sql_statements)}")
    
    # Show first few statements
    print("\nFirst 15 statements:")
    for stmt in sql_statements[:15]:
        print(stmt)

if __name__ == "__main__":
    main()
